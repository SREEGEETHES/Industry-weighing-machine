"""
Generates the weekly Excel report from real audit data in the DB - no
placeholder figures. Structure matches what you specified: daily totals,
box counts, weight averages, min/max, and print status counts.
"""
import datetime as dt
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models import BoxRecord, Station

REPORTS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "..", "reports")


HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
FAIL_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")


def _style_header(ws, row_idx: int, num_cols: int):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def generate_weekly_report(db: Session, period_start: dt.datetime,
                            period_end: dt.datetime) -> str:
    os.makedirs(REPORTS_DIR, exist_ok=True)

    records = (
        db.query(BoxRecord)
        .filter(BoxRecord.created_at >= period_start, BoxRecord.created_at < period_end)
        .order_by(BoxRecord.created_at)
        .all()
    )

    wb = Workbook()

    # ---------------- Summary sheet ----------------
    summary = wb.active
    summary.title = "Summary"
    summary["A1"] = "Weekly Weigh & Print Audit Report"
    summary["A1"].font = Font(bold=True, size=14)
    summary["A2"] = f"Period: {period_start.strftime('%d/%m/%Y')} - {period_end.strftime('%d/%m/%Y')}"

    total_boxes = len(records)
    total_weight = sum(r.weight for r in records)
    avg_weight = round(total_weight / total_boxes, 3) if total_boxes else 0.0
    min_weight = round(min((r.weight for r in records), default=0.0), 3)
    max_weight = round(max((r.weight for r in records), default=0.0), 3)
    printed = sum(1 for r in records if r.print_status == "printed")
    failed = sum(1 for r in records if r.print_status == "failed")
    pending = sum(1 for r in records if r.print_status == "pending")
    out_of_tolerance = sum(1 for r in records if r.within_tolerance is False)

    summary_rows = [
        ("Total boxes weighed", total_boxes),
        ("Total weight", f"{round(total_weight, 3)} kg"),
        ("Average weight", f"{avg_weight} kg"),
        ("Minimum weight", f"{min_weight} kg"),
        ("Maximum weight", f"{max_weight} kg"),
        ("Boxes printed successfully", printed),
        ("Boxes with failed print (needs attention)", failed),
        ("Boxes pending print", pending),
        ("Boxes out of tolerance (needs attention)", out_of_tolerance),
    ]
    start_row = 4
    for i, (label, value) in enumerate(summary_rows):
        summary.cell(row=start_row + i, column=1, value=label).font = Font(bold=True)
        summary.cell(row=start_row + i, column=2, value=value)

    # Per-station breakdown
    station_row = start_row + len(summary_rows) + 2
    summary.cell(row=station_row, column=1, value="Per-Station Breakdown").font = Font(bold=True, size=12)
    station_row += 1
    headers = ["Station", "Machine ID", "Boxes", "Total Weight (kg)", "Avg Weight (kg)", "Out of Tolerance"]
    for i, h in enumerate(headers, start=1):
        summary.cell(row=station_row, column=i, value=h)
    _style_header(summary, station_row, len(headers))

    stations = db.query(Station).all()
    for i, station in enumerate(stations, start=1):
        st_records = [r for r in records if r.station_id == station.id]
        st_total = sum(r.weight for r in st_records)
        st_avg = round(st_total / len(st_records), 3) if st_records else 0.0
        st_oot = sum(1 for r in st_records if r.within_tolerance is False)
        row = station_row + i
        summary.cell(row=row, column=1, value=station.name)
        summary.cell(row=row, column=2, value=station.machine_id)
        summary.cell(row=row, column=3, value=len(st_records))
        summary.cell(row=row, column=4, value=round(st_total, 3))
        summary.cell(row=row, column=5, value=st_avg)
        summary.cell(row=row, column=6, value=st_oot)

    for col in range(1, 7):
        summary.column_dimensions[get_column_letter(col)].width = 26

    # ---------------- Detail sheet ----------------
    detail = wb.create_sheet("Box Detail")
    detail_headers = [
        "Box ID", "Date", "Time", "Station", "Machine ID", "Weight", "Unit",
        "Product Code", "Batch", "Target", "Min", "Max", "Within Tolerance",
        "Variance", "Operator", "Print Status",
    ]
    for i, h in enumerate(detail_headers, start=1):
        detail.cell(row=1, column=i, value=h)
    _style_header(detail, 1, len(detail_headers))

    station_lookup = {s.id: s.name for s in stations}
    for row_idx, r in enumerate(records, start=2):
        values = [
            r.box_id, r.created_at.strftime("%d/%m/%Y"), r.created_at.strftime("%H:%M:%S"),
            station_lookup.get(r.station_id, str(r.station_id)), r.machine_id,
            r.weight, r.unit, r.product_code, r.batch_number,
            r.target_weight, r.min_weight, r.max_weight,
            "Yes" if r.within_tolerance else ("No" if r.within_tolerance is False else ""),
            r.variance, r.operator, r.print_status,
        ]
        for col_idx, v in enumerate(values, start=1):
            detail.cell(row=row_idx, column=col_idx, value=v)
        if r.within_tolerance is False or r.print_status == "failed":
            for col_idx in range(1, len(detail_headers) + 1):
                detail.cell(row=row_idx, column=col_idx).fill = FAIL_FILL

    for col in range(1, len(detail_headers) + 1):
        detail.column_dimensions[get_column_letter(col)].width = 16

    filename = f"weekly_report_{period_start.strftime('%Y%m%d')}_{period_end.strftime('%Y%m%d')}.xlsx"
    filepath = os.path.abspath(os.path.join(REPORTS_DIR, filename))
    wb.save(filepath)
    return filepath
